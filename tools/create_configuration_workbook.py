"""
============================================================
RecruitOS

Module      : Configuration Workbook Generator
Sprint      : 5.5.0.1
Version     : 1.0.0
Author      : Tamilvanan A

Purpose:
Creates the RecruitOS_Configuration.xlsx workbook
with all required sheets, headers, formatting and
sample records.

============================================================
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from config.paths import CONFIGURATION_WORKBOOK
from config.paths import MASTER_DATA_DIR

from config.sheet_names import *


HEADER_FILL = PatternFill(
    fill_type="solid",
    start_color="1F4E78",
    end_color="1F4E78"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF"
)


def format_header(sheet):

    """
    Formats first row.
    """

    for cell in sheet[1]:

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def create_sheet(workbook, sheet_name, headers, sample_rows=None):

    """
    Creates one worksheet.
    """

    sheet = workbook.create_sheet(sheet_name)

    sheet.append(headers)

    format_header(sheet)

    if sample_rows:

        for row in sample_rows:

            sheet.append(row)


def build_workbook():

    MASTER_DATA_DIR.mkdir(
        parents=True,
        exist_ok=True
    )

    workbook = Workbook()

    workbook.remove(workbook.active)

    # -----------------------------------------------------
    # Skills
    # -----------------------------------------------------

    create_sheet(

        workbook,

        SKILLS,

        [

            "Skill",

            "Category",

            "Synonyms",

            "Active"

        ],

        [

            [

                "Python",

                "Programming",

                "Python3,Py",

                "Yes"

            ],

            [

                "SQL",

                "Database",

                "TSQL,PLSQL",

                "Yes"

            ]

        ]

    )

    # -----------------------------------------------------
    # Education
    # -----------------------------------------------------

    create_sheet(

        workbook,

        EDUCATION,

        [

            "Education",

            "Alias",

            "Priority",

            "Active"

        ],

        [

            [

                "Bachelor of Engineering",

                "BE",

                1,

                "Yes"

            ],

            [

                "Master of Engineering",

                "ME",

                2,

                "Yes"

            ]

        ]

    )

    # -----------------------------------------------------
    # Certifications
    # -----------------------------------------------------

    create_sheet(

        workbook,

        CERTIFICATIONS,

        [

            "Certification",

            "Alias",

            "Category",

            "Active"

        ]

    )

    # -----------------------------------------------------
    # Companies
    # -----------------------------------------------------

    create_sheet(

        workbook,

        COMPANIES,

        [

            "Company",

            "Alias",

            "Industry",

            "Active"

        ]

    )

    # -----------------------------------------------------
    # Locations
    # -----------------------------------------------------

    create_sheet(

        workbook,

        LOCATIONS,

        [

            "Country",

            "State",

            "City",

            "Active"

        ]

    )

    # -----------------------------------------------------
    # Domains
    # -----------------------------------------------------

    create_sheet(

        workbook,

        DOMAINS,

        [

            "Domain",

            "Description",

            "Active"

        ]

    )

    # -----------------------------------------------------
    # Languages
    # -----------------------------------------------------

    create_sheet(

        workbook,

        LANGUAGES,

        [

            "Language",

            "Category",

            "Active"

        ]

    )

    # -----------------------------------------------------
    # Roles
    # -----------------------------------------------------

    create_sheet(

        workbook,

        ROLES,

        [

            "Role",

            "Department",

            "Active"

        ]

    )

    # -----------------------------------------------------
    # Industries
    # -----------------------------------------------------

    create_sheet(

        workbook,

        INDUSTRIES,

        [

            "Industry",

            "Description",

            "Active"

        ]

    )

    # -----------------------------------------------------
    # Scoring
    # -----------------------------------------------------

    create_sheet(

        workbook,

        SCORING,

        [

            "Component",

            "Weight",

            "Active",

            "Remarks"

        ],

        [

            [

                "Skill",

                40,

                "Yes",

                "Mandatory Skills"

            ],

            [

                "Experience",

                25,

                "Yes",

                "Experience"

            ],

            [

                "Education",

                15,

                "Yes",

                "Education"

            ],

            [

                "Certification",

                10,

                "Yes",

                "Certification"

            ],

            [

                "Keyword",

                10,

                "Yes",

                "Keyword Matching"

            ]

        ]

    )

    # -----------------------------------------------------
    # Recommendation
    # -----------------------------------------------------

    create_sheet(

        workbook,

        RECOMMENDATION,

        [

            "Minimum Score",

            "Maximum Score",

            "Recommendation"

        ],

        [

            [

                90,

                100,

                "Highly Recommended"

            ],

            [

                75,

                89,

                "Recommended"

            ],

            [

                60,

                74,

                "Review"

            ],

            [

                0,

                59,

                "Not Recommended"

            ]

        ]

    )

    # -----------------------------------------------------
    # Configuration
    # -----------------------------------------------------

    create_sheet(

        workbook,

        CONFIGURATION,

        [

            "Key",

            "Value"

        ],

        [

            [

                "Product",

                "RecruitOS"

            ],

            [

                "Version",

                "1.0.0"

            ],

            [

                "Environment",

                "Development"

            ],

            [

                "AI Enabled",

                "Yes"

            ]

        ]

    )

    workbook.save(CONFIGURATION_WORKBOOK)

    print()

    print("Configuration Workbook Created Successfully")

    print(CONFIGURATION_WORKBOOK)


if __name__ == "__main__":

    build_workbook()