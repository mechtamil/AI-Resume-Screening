"""RecruitOS ranked screening Excel report generation."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from re import sub
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.paths import OUTPUT_DIR


class ExcelReportService:
    """
    Generates RecruitOS ranked screening reports.

    The report is built from the final ProcessingService result:

        {
            "job_description": JobDescription,
            "candidates": [...],
            "match_results": [...],
            "errors": [...],
            "summary": {...}
        }

    This service contains presentation/report logic only.

    It does not:
        - Parse resumes
        - Perform matching
        - Calculate scores
        - Apply recommendation rules
    """

    TITLE_FILL = "1F4E78"
    SECTION_FILL = "D9EAF7"
    HEADER_FILL = "5B9BD5"
    HEADER_FONT = "FFFFFF"

    THIN_BORDER = Border(
        left=Side(style="thin", color="D9E1F2"),
        right=Side(style="thin", color="D9E1F2"),
        top=Side(style="thin", color="D9E1F2"),
        bottom=Side(style="thin", color="D9E1F2"),
    )

    # ========================================================
    # Public API
    # ========================================================

    @classmethod
    def build_report(
        cls,
        analysis_result: dict[str, Any],
    ) -> bytes:
        """
        Build the complete screening report.

        Returns:
            XLSX file content as bytes.
        """

        if not isinstance(
            analysis_result,
            dict,
        ):
            raise TypeError(
                "analysis_result must be a dictionary."
            )

        job = analysis_result.get(
            "job_description"
        )

        if job is None:
            raise ValueError(
                "analysis_result does not contain "
                "job_description."
            )

        matches = list(
            analysis_result.get(
                "match_results"
            )
            or []
        )

        candidates = list(
            analysis_result.get(
                "candidates"
            )
            or []
        )

        errors = list(
            analysis_result.get(
                "errors"
            )
            or []
        )

        summary = dict(
            analysis_result.get(
                "summary"
            )
            or {}
        )

        workbook = Workbook()

        summary_sheet = workbook.active

        summary_sheet.title = (
            "Screening Summary"
        )

        candidate_lookup = (
            cls._candidate_lookup(
                candidates
            )
        )

        cls._write_summary_sheet(
            summary_sheet,
            job,
            matches,
            summary,
        )

        used_titles = {
            summary_sheet.title
        }

        for index, result in enumerate(
            matches,
            start=1,
        ):

            candidate = (
                cls._find_candidate(
                    candidate_lookup,
                    result,
                )
            )

            candidate_name = (
                getattr(
                    result,
                    "candidate_name",
                    "",
                )
                or getattr(
                    result,
                    "source_file",
                    "",
                )
                or "Candidate"
            )

            title = (
                cls._unique_sheet_title(
                    f"{index:02d} "
                    f"{candidate_name}",
                    used_titles,
                )
            )

            used_titles.add(
                title
            )

            sheet = workbook.create_sheet(
                title
            )

            cls._write_candidate_sheet(
                sheet,
                result,
                candidate,
            )

        if errors:

            error_sheet = (
                workbook.create_sheet(
                    "Processing Errors"
                )
            )

            cls._write_errors_sheet(
                error_sheet,
                errors,
            )

        buffer = BytesIO()

        workbook.save(
            buffer
        )

        return buffer.getvalue()

    # ========================================================

    @classmethod
    def save_report(
        cls,
        analysis_result: dict[str, Any],
        output_path: str | Path | None = None,
    ) -> Path:
        """
        Save the screening report to disk.

        Returns:
            Final saved Path.
        """

        job = None

        if isinstance(
            analysis_result,
            dict,
        ):
            job = analysis_result.get(
                "job_description"
            )

        job_title = (
            getattr(
                job,
                "job_title",
                "",
            )
            if job
            else ""
        )

        default_name = (
            cls.default_filename(
                job_title
            )
        )

        path = (
            Path(
                output_path
            )
            if output_path
            else OUTPUT_DIR
            / default_name
        )

        path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        path.write_bytes(
            cls.build_report(
                analysis_result
            )
        )

        return path

    # ========================================================

    @staticmethod
    def default_filename(
        job_title: str = "",
    ) -> str:
        """
        Generate a safe report filename.
        """

        cleaned = sub(
            r"[^A-Za-z0-9._-]+",
            "_",
            str(
                job_title
                or ""
            ).strip(),
        ).strip(
            "._-"
        )

        base = (
            cleaned
            or "RecruitOS_Screening"
        )

        timestamp = (
            datetime.now().strftime(
                "%Y%m%d_%H%M%S"
            )
        )

        return (
            f"{base}_"
            f"Screening_Report_"
            f"{timestamp}.xlsx"
        )

    # ========================================================
    # Summary Sheet
    # ========================================================

    @classmethod
    def _write_summary_sheet(
        cls,
        sheet,
        job: Any,
        matches: list[Any],
        summary: dict[str, Any],
    ) -> None:

        sheet.merge_cells(
            "A1:N1"
        )

        sheet["A1"] = (
            "RecruitOS — Ranked Candidate "
            "Screening Report"
        )

        sheet["A1"].font = Font(
            bold=True,
            color="FFFFFF",
            size=16,
        )

        sheet["A1"].fill = PatternFill(
            "solid",
            fgColor=cls.TITLE_FILL,
        )

        sheet["A1"].alignment = Alignment(
            horizontal="center"
        )

        sheet.row_dimensions[
            1
        ].height = 26

        # ----------------------------------------------------
        # Job Description
        # ----------------------------------------------------

        sheet["A3"] = (
            "Job Description"
        )

        cls._style_section(
            sheet["A3"]
        )

        job_rows = [

            (
                "Job Title",
                getattr(
                    job,
                    "job_title",
                    "",
                ),
            ),

            (
                "Company",
                getattr(
                    job,
                    "company_name",
                    "",
                ),
            ),

            (
                "Location",
                getattr(
                    job,
                    "location",
                    "",
                ),
            ),

            (
                "Employment Type",
                getattr(
                    job,
                    "employment_type",
                    "",
                ),
            ),

            (
                "Experience",
                cls._experience_text(
                    getattr(
                        job,
                        "experience_min",
                        0.0,
                    ),
                    getattr(
                        job,
                        "experience_max",
                        0.0,
                    ),
                ),
            ),

            (
                "Mandatory Skills",
                cls._join(
                    getattr(
                        job,
                        "mandatory_skills",
                        [],
                    )
                ),
            ),

            (
                "Preferred Skills",
                cls._join(
                    getattr(
                        job,
                        "preferred_skills",
                        [],
                    )
                ),
            ),

            (
                "Education",
                cls._join(
                    getattr(
                        job,
                        "education",
                        [],
                    )
                ),
            ),

            (
                "Certifications",
                cls._join(
                    getattr(
                        job,
                        "certifications",
                        [],
                    )
                ),
            ),

            (
                "Keywords",
                cls._join(
                    getattr(
                        job,
                        "keywords",
                        [],
                    )
                ),
            ),
        ]

        for row_number, (
            label,
            value,
        ) in enumerate(
            job_rows,
            start=4,
        ):

            sheet.cell(
                row=row_number,
                column=1,
                value=label,
            ).font = Font(
                bold=True
            )

            sheet.cell(
                row=row_number,
                column=2,
                value=value,
            )

            sheet.merge_cells(
                start_row=row_number,
                start_column=2,
                end_row=row_number,
                end_column=6,
            )

            sheet.cell(
                row=row_number,
                column=2,
            ).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

        # ----------------------------------------------------
        # Screening Statistics
        # ----------------------------------------------------

        stats_row = 15

        sheet.cell(
            stats_row,
            1,
            "Screening Summary",
        )

        cls._style_section(
            sheet.cell(
                stats_row,
                1,
            )
        )

        stats = [

            (
                "Resumes Requested",
                summary.get(
                    "resumes_requested",
                    len(matches),
                ),
            ),

            (
                "Resumes Processed",
                summary.get(
                    "resumes_processed",
                    len(matches),
                ),
            ),

            (
                "Resumes Failed",
                summary.get(
                    "resumes_failed",
                    0,
                ),
            ),

            (
                "Ranked Candidates",
                len(
                    matches
                ),
            ),

            (
                "Shortlisted",
                sum(
                    1
                    for item
                    in matches
                    if bool(
                        getattr(
                            item,
                            "shortlisted",
                            False,
                        )
                    )
                ),
            ),
        ]

        for offset, (
            label,
            value,
        ) in enumerate(
            stats,
            start=1,
        ):

            sheet.cell(
                stats_row
                + offset,
                1,
                label,
            ).font = Font(
                bold=True
            )

            sheet.cell(
                stats_row
                + offset,
                2,
                value,
            )

        # ----------------------------------------------------
        # Ranked Candidates
        # ----------------------------------------------------

        table_start = (
            stats_row
            + len(
                stats
            )
            + 3
        )

        headers = [

            "Rank",
            "Candidate",
            "Email",
            "Phone",
            "Source File",
            "Overall Match %",
            "Recommendation",
            "Shortlisted",
            "Status",
            "Skill Score",
            "Experience Score",
            "Education Score",
            "Certification Score",
            "Keyword Score",

        ]

        for column, header in enumerate(
            headers,
            start=1,
        ):

            cell = sheet.cell(
                table_start,
                column,
                header,
            )

            cls._style_header(
                cell
            )

        for row_index, result in enumerate(
            matches,
            start=table_start + 1,
        ):

            values = [

                getattr(
                    result,
                    "rank",
                    0,
                ),

                getattr(
                    result,
                    "candidate_name",
                    "",
                ),

                getattr(
                    result,
                    "email",
                    "",
                ),

                getattr(
                    result,
                    "phone",
                    "",
                ),

                getattr(
                    result,
                    "source_file",
                    "",
                ),

                float(
                    getattr(
                        result,
                        "overall_match_percentage",
                        0.0,
                    )
                    or 0.0
                ),

                getattr(
                    result,
                    "recommendation",
                    "",
                ),

                (
                    "Yes"
                    if bool(
                        getattr(
                            result,
                            "shortlisted",
                            False,
                        )
                    )
                    else "No"
                ),

                getattr(
                    result,
                    "status",
                    "",
                ),

                float(
                    getattr(
                        result,
                        "skill_score",
                        0.0,
                    )
                    or 0.0
                ),

                float(
                    getattr(
                        result,
                        "experience_score",
                        0.0,
                    )
                    or 0.0
                ),

                float(
                    getattr(
                        result,
                        "education_score",
                        0.0,
                    )
                    or 0.0
                ),

                float(
                    getattr(
                        result,
                        "certification_score",
                        0.0,
                    )
                    or 0.0
                ),

                float(
                    getattr(
                        result,
                        "keyword_score",
                        0.0,
                    )
                    or 0.0
                ),

            ]

            for column, value in enumerate(
                values,
                start=1,
            ):

                cell = sheet.cell(
                    row_index,
                    column,
                    value,
                )

                cell.border = (
                    cls.THIN_BORDER
                )

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

                if column in {
                    6,
                    10,
                    11,
                    12,
                    13,
                    14,
                }:

                    cell.number_format = (
                        "0.00"
                    )

        sheet.freeze_panes = (
            f"A{table_start + 1}"
        )

        sheet.auto_filter.ref = (
            f"A{table_start}:"
            f"N{max(table_start, table_start + len(matches))}"
        )

        cls._set_widths(

            sheet,

            {

                1: 8,
                2: 24,
                3: 28,
                4: 18,
                5: 28,
                6: 16,
                7: 22,
                8: 12,
                9: 14,
                10: 14,
                11: 17,
                12: 15,
                13: 19,
                14: 14,

            },

        )

    # ========================================================
    # Candidate Detail Sheet
    # ========================================================

    @classmethod
    def _write_candidate_sheet(
        cls,
        sheet,
        result: Any,
        candidate: Any | None,
    ) -> None:

        candidate_name = (

            getattr(
                result,
                "candidate_name",
                "",
            )

            or getattr(
                result,
                "source_file",
                "",
            )

            or "Candidate"

        )

        sheet.merge_cells(
            "A1:F1"
        )

        sheet["A1"] = (
            "Candidate Detail — "
            f"{candidate_name}"
        )

        sheet["A1"].font = Font(
            bold=True,
            color="FFFFFF",
            size=15,
        )

        sheet["A1"].fill = PatternFill(
            "solid",
            fgColor=cls.TITLE_FILL,
        )

        sheet["A1"].alignment = Alignment(
            horizontal="center"
        )

        row = 3

        row = cls._write_key_value_section(

            sheet,

            row,

            "Candidate Information",

            [

                (
                    "Name",
                    getattr(
                        result,
                        "candidate_name",
                        "",
                    )
                    or getattr(
                        candidate,
                        "full_name",
                        "",
                    ),
                ),

                (
                    "Email",
                    getattr(
                        result,
                        "email",
                        "",
                    )
                    or getattr(
                        candidate,
                        "email",
                        "",
                    ),
                ),

                (
                    "Phone",
                    getattr(
                        result,
                        "phone",
                        "",
                    )
                    or getattr(
                        candidate,
                        "phone",
                        "",
                    ),
                ),

                (
                    "Location",
                    getattr(
                        candidate,
                        "location",
                        "",
                    ),
                ),

                (
                    "Source File",
                    getattr(
                        result,
                        "source_file",
                        "",
                    )
                    or getattr(
                        candidate,
                        "source_file",
                        "",
                    ),
                ),

                (
                    "Candidate Experience",
                    float(
                        getattr(
                            result,
                            "candidate_experience",
                            0.0,
                        )
                        or 0.0
                    ),
                ),

                (
                    "Education",
                    cls._join(
                        getattr(
                            candidate,
                            "education",
                            [],
                        )
                    ),
                ),

                (
                    "Certifications",
                    cls._join(
                        getattr(
                            candidate,
                            "certifications",
                            [],
                        )
                    ),
                ),

                (
                    "Technical Skills",
                    cls._join(
                        getattr(
                            candidate,
                            "technical_skills",
                            [],
                        )
                    ),
                ),

            ],

        )

        row = cls._write_key_value_section(

            sheet,

            row + 1,

            "Match Summary",

            [

                (
                    "Rank",
                    getattr(
                        result,
                        "rank",
                        0,
                    ),
                ),

                (
                    "Overall Match %",
                    float(
                        getattr(
                            result,
                            "overall_match_percentage",
                            0.0,
                        )
                        or 0.0
                    ),
                ),

                (
                    "Recommendation",
                    getattr(
                        result,
                        "recommendation",
                        "",
                    ),
                ),

                (
                    "Shortlisted",
                    (
                        "Yes"
                        if bool(
                            getattr(
                                result,
                                "shortlisted",
                                False,
                            )
                        )
                        else "No"
                    ),
                ),

                (
                    "Status",
                    getattr(
                        result,
                        "status",
                        "",
                    ),
                ),

                (
                    "Required Experience",
                    float(
                        getattr(
                            result,
                            "required_experience",
                            0.0,
                        )
                        or 0.0
                    ),
                ),

                (
                    "Maximum Experience",
                    float(
                        getattr(
                            result,
                            "maximum_experience",
                            0.0,
                        )
                        or 0.0
                    ),
                ),

                (
                    "Experience Match",
                    (
                        "Yes"
                        if bool(
                            getattr(
                                result,
                                "experience_match",
                                False,
                            )
                        )
                        else "No"
                    ),
                ),

                (
                    "Education Match",
                    (
                        "Yes"
                        if bool(
                            getattr(
                                result,
                                "education_match",
                                False,
                            )
                        )
                        else "No"
                    ),
                ),

                (
                    "Certification Match",
                    (
                        "Yes"
                        if bool(
                            getattr(
                                result,
                                "certification_match",
                                False,
                            )
                        )
                        else "No"
                    ),
                ),

            ],

            percentage_labels={
                "Overall Match %"
            },

        )

        row = cls._write_key_value_section(

            sheet,

            row + 1,

            "Component Scores",

            [

                (
                    "Skill Score",
                    float(
                        getattr(
                            result,
                            "skill_score",
                            0.0,
                        )
                        or 0.0
                    ),
                ),

                (
                    "Experience Score",
                    float(
                        getattr(
                            result,
                            "experience_score",
                            0.0,
                        )
                        or 0.0
                    ),
                ),

                (
                    "Education Score",
                    float(
                        getattr(
                            result,
                            "education_score",
                            0.0,
                        )
                        or 0.0
                    ),
                ),

                (
                    "Certification Score",
                    float(
                        getattr(
                            result,
                            "certification_score",
                            0.0,
                        )
                        or 0.0
                    ),
                ),

                (
                    "Keyword Score",
                    float(
                        getattr(
                            result,
                            "keyword_score",
                            0.0,
                        )
                        or 0.0
                    ),
                ),

            ],

            percentage_labels={

                "Skill Score",
                "Experience Score",
                "Education Score",
                "Certification Score",
                "Keyword Score",

            },

        )

        # ----------------------------------------------------
        # Weighted Score Breakdown
        # ----------------------------------------------------

        breakdown = dict(

            getattr(
                result,
                "weighted_score_breakdown",
                {},
            )

            or {}

        )

        if breakdown:

            sheet.cell(
                row,
                1,
                "Weighted Score Breakdown",
            )

            cls._style_section(
                sheet.cell(
                    row,
                    1,
                )
            )

            row += 1

            for header_col, header in enumerate(

                [
                    "Component",
                    "Weighted Contribution",
                ],

                start=1,

            ):

                cls._style_header(

                    sheet.cell(
                        row,
                        header_col,
                        header,
                    )

                )

            row += 1

            for component, contribution in (

                breakdown.items()

            ):

                sheet.cell(
                    row,
                    1,
                    str(
                        component
                    ),
                )

                sheet.cell(
                    row,
                    2,
                    float(
                        contribution
                        or 0.0
                    ),
                ).number_format = (
                    "0.00"
                )

                row += 1

        # ----------------------------------------------------
        # Match Details
        # ----------------------------------------------------

        list_sections = [

            (
                "Matched Mandatory Skills",
                getattr(
                    result,
                    "matched_skills",
                    [],
                ),
            ),

            (
                "Missing Mandatory Skills",
                getattr(
                    result,
                    "missing_skills",
                    [],
                ),
            ),

            (
                "Matched Preferred Skills",
                getattr(
                    result,
                    "matched_preferred_skills",
                    [],
                ),
            ),

            (
                "Missing Preferred Skills",
                getattr(
                    result,
                    "missing_preferred_skills",
                    [],
                ),
            ),

            (
                "Additional Skills",
                getattr(
                    result,
                    "additional_skills",
                    [],
                ),
            ),

            (
                "Matched Certifications",
                getattr(
                    result,
                    "matched_certifications",
                    [],
                ),
            ),

            (
                "Missing Certifications",
                getattr(
                    result,
                    "missing_certifications",
                    [],
                ),
            ),

            (
                "Matched Keywords",
                getattr(
                    result,
                    "matched_keyword_values",
                    [],
                ),
            ),

            (
                "Missing Keywords",
                getattr(
                    result,
                    "missing_keyword_values",
                    [],
                ),
            ),

            (
                "Remarks",
                getattr(
                    result,
                    "remarks",
                    [],
                ),
            ),

        ]

        for title, values in list_sections:

            row = cls._write_key_value_section(

                sheet,

                row + 1,

                title,

                [

                    (
                        title,
                        cls._join(
                            values
                        ),
                    )

                ],

            )

        cls._set_widths(

            sheet,

            {

                1: 30,
                2: 75,
                3: 18,
                4: 18,
                5: 18,
                6: 18,

            },

        )

        for row_cells in sheet.iter_rows():

            for cell in row_cells:

                cell.alignment = Alignment(

                    vertical="top",
                    wrap_text=True,

                )

    # ========================================================
    # Errors Sheet
    # ========================================================

    @classmethod
    def _write_errors_sheet(
        cls,
        sheet,
        errors: list[dict[str, Any]],
    ) -> None:

        sheet["A1"] = "File"
        sheet["B1"] = "Error"

        cls._style_header(
            sheet["A1"]
        )

        cls._style_header(
            sheet["B1"]
        )

        for row, error in enumerate(
            errors,
            start=2,
        ):

            sheet.cell(
                row,
                1,
                str(
                    error.get(
                        "file",
                        "",
                    )
                ),
            )

            sheet.cell(
                row,
                2,
                str(
                    error.get(
                        "error",
                        "",
                    )
                ),
            )

            sheet.cell(
                row,
                2,
            ).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

        cls._set_widths(

            sheet,

            {
                1: 32,
                2: 90,
            },

        )

        sheet.freeze_panes = (
            "A2"
        )

        sheet.auto_filter.ref = (

            f"A1:B"
            f"{max(1, len(errors) + 1)}"

        )

    # ========================================================
    # Generic Section Writer
    # ========================================================

    @classmethod
    def _write_key_value_section(
        cls,
        sheet,
        start_row: int,
        title: str,
        items: list[
            tuple[
                str,
                Any,
            ]
        ],
        percentage_labels: set[str]
        | None = None,
    ) -> int:

        percentage_labels = (
            percentage_labels
            or set()
        )

        sheet.cell(
            start_row,
            1,
            title,
        )

        cls._style_section(

            sheet.cell(
                start_row,
                1,
            )

        )

        row = (
            start_row
            + 1
        )

        for label, value in items:

            label_cell = (
                sheet.cell(
                    row,
                    1,
                    label,
                )
            )

            value_cell = (
                sheet.cell(
                    row,
                    2,
                    value,
                )
            )

            label_cell.font = Font(
                bold=True
            )

            value_cell.alignment = Alignment(

                wrap_text=True,
                vertical="top",

            )

            if (

                label
                in percentage_labels

                and isinstance(
                    value,
                    (
                        int,
                        float,
                    ),
                )

            ):

                value_cell.number_format = (
                    "0.00"
                )

            row += 1

        return row

    # ========================================================
    # Formatting Helpers
    # ========================================================

    @classmethod
    def _style_section(
        cls,
        cell,
    ) -> None:

        cell.font = Font(
            bold=True
        )

        cell.fill = PatternFill(

            "solid",

            fgColor=(
                cls.SECTION_FILL
            ),

        )

    # ========================================================

    @classmethod
    def _style_header(
        cls,
        cell,
    ) -> None:

        cell.font = Font(

            bold=True,

            color=(
                cls.HEADER_FONT
            ),

        )

        cell.fill = PatternFill(

            "solid",

            fgColor=(
                cls.HEADER_FILL
            ),

        )

        cell.border = (
            cls.THIN_BORDER
        )

        cell.alignment = Alignment(

            horizontal="center",

            vertical="center",

            wrap_text=True,

        )

    # ========================================================

    @staticmethod
    def _set_widths(
        sheet,
        widths: dict[int, float],
    ) -> None:

        for column, width in (
            widths.items()
        ):

            sheet.column_dimensions[
                get_column_letter(
                    column
                )
            ].width = width

    # ========================================================
    # Utility Helpers
    # ========================================================

    @staticmethod
    def _join(
        values: Any,
    ) -> str:

        if values is None:

            return ""

        if isinstance(
            values,
            str,
        ):

            return values.strip()

        try:

            return ", ".join(

                str(
                    item
                ).strip()

                for item in values

                if str(
                    item
                ).strip()

            )

        except TypeError:

            return str(
                values
            )

    # ========================================================

    @staticmethod
    def _experience_text(
        minimum: Any,
        maximum: Any,
    ) -> str:

        try:

            minimum_value = float(
                minimum
                or 0.0
            )

            maximum_value = float(
                maximum
                or 0.0
            )

        except (
            TypeError,
            ValueError,
        ):

            return ""

        if maximum_value:

            return (

                f"{minimum_value:g}"
                f"–"
                f"{maximum_value:g} years"

            )

        if minimum_value:

            return (
                f"{minimum_value:g}+ years"
            )

        return (
            "Not specified"
        )

    # ========================================================

    @staticmethod
    def _candidate_lookup(
        candidates: list[Any],
    ) -> dict[str, Any]:

        lookup: dict[
            str,
            Any,
        ] = {}

        for candidate in candidates:

            source_file = str(

                getattr(
                    candidate,
                    "source_file",
                    "",
                )

                or ""

            ).casefold()

            email = str(

                getattr(
                    candidate,
                    "email",
                    "",
                )

                or ""

            ).casefold()

            name = str(

                getattr(
                    candidate,
                    "full_name",
                    "",
                )

                or ""

            ).casefold()

            for key in (

                source_file,
                email,
                name,

            ):

                if key:

                    lookup[
                        key
                    ] = candidate

        return lookup

    # ========================================================

    @staticmethod
    def _find_candidate(
        lookup: dict[str, Any],
        result: Any,
    ) -> Any | None:

        for raw in (

            getattr(
                result,
                "source_file",
                "",
            ),

            getattr(
                result,
                "email",
                "",
            ),

            getattr(
                result,
                "candidate_name",
                "",
            ),

        ):

            key = str(
                raw
                or ""
            ).casefold()

            if (
                key
                and key in lookup
            ):

                return lookup[
                    key
                ]

        return None

    # ========================================================

    @staticmethod
    def _unique_sheet_title(
        raw_title: str,
        used_titles: set[str],
    ) -> str:
        """
        Create valid unique Excel worksheet title.

        Excel limitations:
            maximum 31 characters

        Invalid:
            \\ / * ? : [ ]
        """

        cleaned = sub(

            r"[\\/*?:\[\]]+",

            " ",

            str(
                raw_title
                or ""
            ),

        ).strip()

        if not cleaned:

            cleaned = (
                "Candidate"
            )

        base = cleaned[
            :31
        ]

        if base not in used_titles:

            return base

        counter = 2

        while True:

            suffix = (
                f" {counter}"
            )

            candidate = (

                f"{base[:31 - len(suffix)]}"
                f"{suffix}"

            )

            if candidate not in used_titles:

                return candidate

            counter += 1