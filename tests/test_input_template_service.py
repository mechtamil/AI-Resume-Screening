"""Tests for recruiter-downloadable Excel input templates."""
from __future__ import annotations

import unittest
from io import BytesIO

from openpyxl import load_workbook

from services.input_template_service import InputTemplateService


class InputTemplateServiceTests(unittest.TestCase):
    def test_job_description_template_contains_required_structure(self):
        data = InputTemplateService.build_job_description_template()
        workbook = load_workbook(BytesIO(data))
        self.assertIn("Job Description", workbook.sheetnames)
        self.assertIn("Instructions", workbook.sheetnames)
        sheet = workbook["Job Description"]
        self.assertEqual([cell.value for cell in sheet[1]], ["Field", "Value", "Importance", "Guidance"])
        fields = [sheet.cell(row=index, column=1).value for index in range(2, sheet.max_row + 1)]
        self.assertIn("Job Title", fields)
        self.assertIn("Mandatory Skills", fields)
        self.assertIn("Preferred Skills", fields)

    def test_skill_template_contains_requirement_type_validation(self):
        data = InputTemplateService.build_skill_list_template()
        workbook = load_workbook(BytesIO(data))
        sheet = workbook["Skills"]
        self.assertEqual(
            [cell.value for cell in sheet[1]],
            ["Skill", "Requirement Type", "Priority", "Notes"],
        )
        validations = list(sheet.data_validations.dataValidation)
        self.assertGreaterEqual(len(validations), 2)
        self.assertTrue(any("Mandatory,Preferred" in str(item.formula1) for item in validations))


if __name__ == "__main__":
    unittest.main()
