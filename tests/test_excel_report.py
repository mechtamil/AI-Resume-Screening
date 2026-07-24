"""Tests for RecruitOS Excel screening report generation."""

from __future__ import annotations

import tempfile
import unittest

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from JD.jd_model import (
    JobDescription,
)

from models.candidate import (
    Candidate,
)

from models.match_result import (
    MatchResult,
)

from reports.excel_report import (
    ExcelReportService,
)


class ExcelReportServiceTest(
    unittest.TestCase
):

    def setUp(
        self,
    ) -> None:

        self.job = JobDescription(

            job_title=(
                "Technical Documentation Engineer"
            ),

            company_name=(
                "Example Company"
            ),

            location=(
                "Chennai"
            ),

            experience_min=4,

            experience_max=8,

            mandatory_skills=[
                "DITA",
                "XML",
            ],

            preferred_skills=[
                "Python",
            ],

            education=[
                "Bachelor of Engineering",
            ],

            certifications=[
                "Example Certification",
            ],

            keywords=[
                "technical documentation",
            ],

        )

        self.candidate = Candidate(

            full_name=(
                "Test Candidate"
            ),

            email=(
                "candidate@example.com"
            ),

            phone=(
                "9999999999"
            ),

            location=(
                "Chennai"
            ),

            total_experience=6,

            education=[
                "Bachelor of Engineering",
            ],

            certifications=[
                "Example Certification",
            ],

            technical_skills=[

                "DITA",
                "XML",
                "Python",

            ],

            source_file=(
                "test_resume.pdf"
            ),

        )

        self.match = MatchResult(

            candidate_name=(
                "Test Candidate"
            ),

            email=(
                "candidate@example.com"
            ),

            phone=(
                "9999999999"
            ),

            source_file=(
                "test_resume.pdf"
            ),

            job_title=(
                self.job.job_title
            ),

            matched_skills=[
                "DITA",
                "XML",
            ],

            matched_preferred_skills=[
                "Python",
            ],

            matched_certifications=[
                "Example Certification",
            ],

            matched_keyword_values=[
                "technical documentation",
            ],

            skill_score=100.0,

            experience_score=100.0,

            education_score=100.0,

            certification_score=100.0,

            keyword_score=100.0,

            weighted_score_breakdown={

                "Skill": 40.0,

                "Experience": 25.0,

            },

            overall_match_percentage=100.0,

            rank=1,

            recommendation=(
                "Highly Recommended"
            ),

            shortlisted=True,

            status=(
                "Shortlisted"
            ),

            candidate_experience=6,

            required_experience=4,

            maximum_experience=8,

            experience_match=True,

            education_match=True,

            certification_match=True,

            remarks=[
                "Strong overall match.",
            ],

        )

        self.analysis_result = {

            "job_description":
                self.job,

            "candidates": [
                self.candidate
            ],

            "match_results": [
                self.match
            ],

            "errors": [

                {

                    "file":
                        "bad_resume.pdf",

                    "error":
                        "Unable to read document",

                }

            ],

            "summary": {

                "resumes_requested":
                    2,

                "resumes_processed":
                    1,

                "resumes_failed":
                    1,

            },

        }

    # ========================================================

    def test_build_report_creates_expected_sheets_and_values(
        self,
    ) -> None:

        report_bytes = (
            ExcelReportService.build_report(
                self.analysis_result
            )
        )

        self.assertGreater(
            len(
                report_bytes
            ),
            1000,
        )

        workbook = load_workbook(

            BytesIO(
                report_bytes
            ),

            data_only=False,

        )

        self.assertIn(

            "Screening Summary",

            workbook.sheetnames,

        )

        self.assertIn(

            "Processing Errors",

            workbook.sheetnames,

        )

        self.assertEqual(

            len(
                workbook.sheetnames
            ),

            3,

        )

        summary = workbook[
            "Screening Summary"
        ]

        self.assertEqual(

            summary[
                "A1"
            ].value,

            (
                "RecruitOS — Ranked Candidate "
                "Screening Report"
            ),

        )

        self.assertEqual(

            summary[
                "B4"
            ].value,

            self.job.job_title,

        )

        summary_values = [

            cell.value

            for row
            in summary.iter_rows()

            for cell
            in row

        ]

        self.assertIn(

            "Test Candidate",

            summary_values,

        )

        self.assertIn(

            "Highly Recommended",

            summary_values,

        )

        detail_sheet = workbook[

            workbook.sheetnames[
                1
            ]

        ]

        detail_values = [

            cell.value

            for row
            in detail_sheet.iter_rows()

            for cell
            in row

        ]

        self.assertIn(

            "Test Candidate",

            detail_values,

        )

        self.assertIn(

            "Bachelor of Engineering",

            detail_values,

        )

        self.assertIn(

            "DITA, XML, Python",

            detail_values,

        )

        errors = workbook[
            "Processing Errors"
        ]

        self.assertEqual(

            errors[
                "A2"
            ].value,

            "bad_resume.pdf",

        )

        self.assertEqual(

            errors[
                "B2"
            ].value,

            "Unable to read document",

        )

    # ========================================================

    def test_save_report_writes_valid_xlsx(
        self,
    ) -> None:

        with tempfile.TemporaryDirectory() as directory:

            output_path = (

                Path(
                    directory
                )

                / "report.xlsx"

            )

            final_path = (

                ExcelReportService.save_report(

                    self.analysis_result,

                    output_path,

                )

            )

            self.assertEqual(

                final_path,

                output_path,

            )

            self.assertTrue(

                output_path.exists()

            )

            workbook = load_workbook(
                output_path
            )

            self.assertIn(

                "Screening Summary",

                workbook.sheetnames,

            )

    # ========================================================

    def test_default_filename_is_safe(
        self,
    ) -> None:

        filename = (

            ExcelReportService.default_filename(

                "JD / Test: Engineer?"

            )

        )

        self.assertTrue(

            filename.endswith(
                ".xlsx"
            )

        )

        self.assertNotIn(
            "/",
            filename,
        )

        self.assertNotIn(
            ":",
            filename,
        )

        self.assertNotIn(
            "?",
            filename,
        )

    # ========================================================

    def test_missing_job_description_is_rejected(
        self,
    ) -> None:

        with self.assertRaises(
            ValueError
        ):

            ExcelReportService.build_report(

                {
                    "match_results": []
                }

            )


if __name__ == "__main__":

    unittest.main()