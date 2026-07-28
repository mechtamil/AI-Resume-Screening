"""Admin user provisioning, Excel import, and credential export tests."""
from __future__ import annotations

import tempfile
import unittest
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from services.auth_service import AuthService
from services.authorization_service import READER, TENANT_ADMIN, USER
from services.user_management_service import UserManagementService
from tests.security_test_utils import (
    TEST_LOCATION,
    TEST_TEMPORARY_PASSWORD,
    create_context,
    create_owner_context,
)


class UserManagementServiceTests(unittest.TestCase):
    def test_single_user_creation_uses_seven_day_reset_credential(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "users.db"
            owner = create_owner_context(path)
            before = datetime.now(timezone.utc)
            created = UserManagementService.create_user(
                owner,
                employee_user_id="6276",
                full_name="Tamilvanan Arumugam",
                email="tamilvanan@example.com",
                role=USER,
                country_location=TEST_LOCATION,
                temporary_password=TEST_TEMPORARY_PASSWORD,
                time_zone="",
                database_path=path,
            )
            after = datetime.now(timezone.utc)

            self.assertEqual(created["temporary_password"], TEST_TEMPORARY_PASSWORD)
            self.assertTrue(created["must_change_password"])
            self.assertEqual(created["account_status"], "RESET_REQUIRED")
            expiry = datetime.fromisoformat(str(created["temporary_password_expires_at"]))
            self.assertGreaterEqual(expiry, before + timedelta(days=7) - timedelta(seconds=2))
            self.assertLessEqual(expiry, after + timedelta(days=7) + timedelta(seconds=2))

            context, _ = AuthService.authenticate(
                user_id="6276",
                password=TEST_TEMPORARY_PASSWORD,
                database_path=path,
            )
            self.assertTrue(context.must_change_password)

    def test_disabled_account_is_provisioned_but_cannot_sign_in(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "users.db"
            owner = create_owner_context(path)
            created = UserManagementService.create_user(
                owner,
                employee_user_id="DISABLED-001",
                full_name="Disabled User",
                email="disabled.user@example.com",
                role=USER,
                country_location=TEST_LOCATION,
                account_status="DISABLED",
                temporary_password=TEST_TEMPORARY_PASSWORD,
                database_path=path,
            )
            self.assertEqual(created["account_status"], "DISABLED")
            with self.assertRaisesRegex(PermissionError, "not active|not available"):
                AuthService.authenticate(
                    user_id="DISABLED-001",
                    password=TEST_TEMPORARY_PASSWORD,
                    database_path=path,
                )

    def test_tenant_admin_can_manage_only_user_and_reader_in_same_location(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "users.db"
            tenant_admin = create_context(
                path,
                "tenant.admin@example.com",
                "Tenant Admin",
                role=TENANT_ADMIN,
                country_location=TEST_LOCATION,
            )
            created = UserManagementService.create_user(
                tenant_admin,
                employee_user_id="LOCAL-USER",
                full_name="Local User",
                email="local.user@example.com",
                role=READER,
                country_location=TEST_LOCATION,
                temporary_password=TEST_TEMPORARY_PASSWORD,
                database_path=path,
            )
            self.assertEqual(created["role"], READER)

            with self.assertRaisesRegex(PermissionError, "assigned country/location"):
                UserManagementService.create_user(
                    tenant_admin,
                    employee_user_id="OTHER-USER",
                    full_name="Other User",
                    email="other.user@example.com",
                    role=USER,
                    country_location="France - Paris",
                    temporary_password=TEST_TEMPORARY_PASSWORD,
                    database_path=path,
                )

    def test_excel_template_preview_import_and_one_time_credentials(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "users.db"
            owner = create_owner_context(path)

            template = UserManagementService.build_import_template()
            workbook = load_workbook(BytesIO(template))
            self.assertIn("Users", workbook.sheetnames)
            self.assertIn("Instructions", workbook.sheetnames)
            sheet = workbook["Users"]
            headers = [cell.value for cell in sheet[1]]
            self.assertEqual(headers, UserManagementService.IMPORT_COLUMNS)

            values = {
                "User ID": "IMPORT-100",
                "Full Name": "Imported User",
                "Email": "imported.user@example.com",
                "Role": "USER",
                "Country or Location": TEST_LOCATION,
                "Account Status": "RESET_REQUIRED",
                "Temporary Password": "Import@123",
            }
            sheet.append([values.get(header, "") for header in headers])
            buffer = BytesIO()
            workbook.save(buffer)

            preview = UserManagementService.preview_import(
                owner,
                buffer.getvalue(),
                database_path=path,
            )
            self.assertEqual(preview["summary"], {"total": 1, "valid": 1, "invalid": 0})

            result = UserManagementService.commit_import(
                owner,
                preview,
                filename="users.xlsx",
                database_path=path,
            )
            self.assertEqual(result["created"], 1)
            self.assertEqual(result["failed"], 0)
            self.assertEqual(result["credentials"][0]["temporary_password"], "Import@123")

            context, _ = AuthService.authenticate(
                user_id="IMPORT-100",
                password="Import@123",
                database_path=path,
            )
            self.assertTrue(context.must_change_password)

    def test_import_detects_duplicates_and_generates_password_when_blank(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "users.db"
            owner = create_owner_context(path)
            workbook = load_workbook(BytesIO(UserManagementService.build_import_template()))
            sheet = workbook["Users"]
            headers = [cell.value for cell in sheet[1]]

            first = {
                "User ID": "BULK-001",
                "Full Name": "Bulk One",
                "Email": "bulk.one@example.com",
                "Role": "USER",
                "Country or Location": TEST_LOCATION,
                "Account Status": "RESET_REQUIRED",
            }
            duplicate = {
                "User ID": "bulk-001",
                "Full Name": "Bulk Duplicate",
                "Email": "bulk.duplicate@example.com",
                "Role": "USER",
                "Country or Location": TEST_LOCATION,
                "Account Status": "RESET_REQUIRED",
            }
            sheet.append([first.get(header, "") for header in headers])
            sheet.append([duplicate.get(header, "") for header in headers])
            buffer = BytesIO()
            workbook.save(buffer)

            preview = UserManagementService.preview_import(owner, buffer.getvalue(), database_path=path)
            self.assertEqual(preview["summary"]["valid"], 1)
            self.assertEqual(preview["summary"]["invalid"], 1)
            self.assertIn("Duplicate User ID", " ".join(preview["rows"][1]["errors"]))

            result = UserManagementService.commit_import(
                owner,
                preview,
                filename="bulk.xlsx",
                database_path=path,
            )
            self.assertEqual(result["created"], 1)
            self.assertNotIn("temporary_password", result["results"][0])
            generated = result["credentials"][0]["temporary_password"]
            self.assertGreaterEqual(len(generated), 12)

    def test_user_access_master_never_contains_plaintext_password(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "users.db"
            owner = create_owner_context(path)
            UserManagementService.create_user(
                owner,
                employee_user_id="MASTER-001",
                full_name="Master User",
                email="master.user@example.com",
                role=USER,
                country_location=TEST_LOCATION,
                temporary_password="NeverExportThis@123",
                database_path=path,
            )

            access_bytes = UserManagementService.export_user_access_master(owner, path)
            workbook = load_workbook(BytesIO(access_bytes), data_only=True)
            sheet = workbook["User Access Master"]
            headers = [cell.value for cell in sheet[1]]
            all_values = [cell.value for row in sheet.iter_rows() for cell in row]
            self.assertNotIn("Temporary Password", headers)
            self.assertNotIn("NeverExportThis@123", all_values)
            self.assertIn("Temporary Password Expiry", headers)

    def test_temporary_credential_export_contains_only_supplied_one_time_values(self):
        credentials = [
            {
                "user_id": "6276",
                "full_name": "Tamilvanan",
                "email": "tamilvanan@example.com",
                "role": "USER",
                "country_location": TEST_LOCATION,
                "temporary_password": "OneTime@123",
                "temporary_password_expires_at": "2026-08-04T12:00:00+00:00",
            }
        ]
        workbook = load_workbook(
            BytesIO(UserManagementService.export_temporary_credentials(credentials)),
            data_only=True,
        )
        values = [cell.value for row in workbook.active.iter_rows() for cell in row]
        self.assertIn("OneTime@123", values)
        self.assertIn("First Login Reset Required", values)


if __name__ == "__main__":
    unittest.main()
