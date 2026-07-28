"""Generate recruiter-friendly Excel templates for structured RecruitOS inputs."""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


class InputTemplateService:
    """Build in-memory Excel templates; no business values are hardcoded."""

    HEADER_FILL = "043962"
    SECTION_FILL = "D9EAF7"
    WHITE = "FFFFFF"

    JD_FIELDS = (
        ("Job Title", "Required", "Exact position name"),
        ("Company Name", "Optional", "Client or business name"),
        ("Location", "Optional", "Work location or remote arrangement"),
        ("Employment Type", "Optional", "Permanent, contract, internship, etc."),
        ("Experience", "Recommended", "Example: 3 to 5 years"),
        ("Education", "Recommended", "Use one item per line"),
        ("Mandatory Skills", "Required", "Use one skill per line"),
        ("Preferred Skills", "Optional", "Use one skill per line"),
        ("Certifications", "Optional", "Use one certification per line"),
        ("Responsibilities", "Recommended", "Use one responsibility per line"),
        ("Keywords", "Optional", "Use one keyword per line"),
        ("Notes", "Optional", "Any additional screening guidance"),
    )

    @classmethod
    def build_job_description_template(cls) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Job Description"
        sheet.append(["Field", "Value", "Importance", "Guidance"])
        cls._style_header(sheet, 1)
        for field, importance, guidance in cls.JD_FIELDS:
            sheet.append([field, "", importance, guidance])

        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 60
        sheet.column_dimensions["C"].width = 16
        sheet.column_dimensions["D"].width = 48
        for row in sheet.iter_rows(min_row=2):
            row[1].alignment = Alignment(wrap_text=True, vertical="top")
            row[3].alignment = Alignment(wrap_text=True, vertical="top")

        instructions = workbook.create_sheet("Instructions")
        instructions.append(["RecruitOS Job Description Template"])
        instructions.append([
            "Complete the Value column. For skills, education, certifications, "
            "responsibilities and keywords, enter one item per line. Structured "
            "inputs improve extraction completeness and matching accuracy."
        ])
        instructions.append([
            "Do not add candidate-specific personal characteristics or protected attributes."
        ])
        instructions.column_dimensions["A"].width = 110
        instructions["A1"].font = Font(bold=True, size=14, color=cls.WHITE)
        instructions["A1"].fill = PatternFill("solid", fgColor=cls.HEADER_FILL)
        instructions["A2"].alignment = Alignment(wrap_text=True)
        instructions["A3"].alignment = Alignment(wrap_text=True)

        return cls._save(workbook)

    @classmethod
    def build_skill_list_template(cls) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Skills"
        sheet.append(["Skill", "Requirement Type", "Priority", "Notes"])
        cls._style_header(sheet, 1)
        for _ in range(25):
            sheet.append(["", "Mandatory", "Medium", ""])

        requirement_validation = DataValidation(
            type="list",
            formula1='"Mandatory,Preferred"',
            allow_blank=False,
        )
        priority_validation = DataValidation(
            type="list",
            formula1='"High,Medium,Low"',
            allow_blank=True,
        )
        sheet.add_data_validation(requirement_validation)
        sheet.add_data_validation(priority_validation)
        requirement_validation.add("B2:B500")
        priority_validation.add("C2:C500")

        sheet.freeze_panes = "A2"
        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 14
        sheet.column_dimensions["D"].width = 55

        instructions = workbook.create_sheet("Instructions")
        instructions.append(["RecruitOS Supplemental Skill List"])
        instructions.append([
            "Add one skill per row. Mandatory skills affect required-skill coverage; "
            "Preferred skills are scored separately. The list supplements the JD and "
            "does not replace it."
        ])
        instructions.column_dimensions["A"].width = 110
        instructions["A1"].font = Font(bold=True, size=14, color=cls.WHITE)
        instructions["A1"].fill = PatternFill("solid", fgColor=cls.HEADER_FILL)
        instructions["A2"].alignment = Alignment(wrap_text=True)
        return cls._save(workbook)

    @classmethod
    def _style_header(cls, sheet, row_number: int) -> None:
        for cell in sheet[row_number]:
            cell.font = Font(bold=True, color=cls.WHITE)
            cell.fill = PatternFill("solid", fgColor=cls.HEADER_FILL)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def _save(workbook: Workbook) -> bytes:
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
