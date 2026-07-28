"""Admin-controlled RecruitOS user provisioning, import, and access exports."""
from __future__ import annotations

import re
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config.brand import ALTEN_BLUE, ALTEN_NAVY
from config.settings import AUTH_TEMP_PASSWORD_EXPIRY_DAYS
from database.user_repository import UserRepository
from models.security_context import SecurityContext
from services.auth_service import AuthService
from services.authorization_service import (
    GLOBAL_ADMIN,
    PERMISSION_ACCESS_MASTER,
    PERMISSION_USER_MANAGE_TENANT,
    SYSTEM_OWNER,
    TENANT_ADMIN,
    AuthorizationService,
    ROLE_LABELS,
)
from services.password_service import PasswordService


class UserManagementService:
    """Provision users without exposing private screening data to administrators."""

    IMPORT_COLUMNS = [
        "User ID",
        "Full Name",
        "Email",
        "Role",
        "Country or Location",
        "Account Status",
        "Time Zone",
        "Department",
        "Business Unit",
        "Manager User ID",
        "Valid From",
        "Valid Until",
        "Temporary Password",
        "Remarks",
    ]
    REQUIRED_COLUMNS = {
        "User ID",
        "Full Name",
        "Email",
        "Role",
        "Country or Location",
        "Account Status",
    }

    @classmethod
    def create_user(
        cls,
        context: SecurityContext,
        *,
        employee_user_id: str,
        full_name: str,
        email: str,
        role: str,
        country_location: str,
        temporary_password: str = "",
        generate_password: bool = False,
        account_status: str = "RESET_REQUIRED",
        time_zone: str = "",
        department: str = "",
        business_unit: str = "",
        manager_user_id: str = "",
        valid_from: str = "",
        valid_until: str = "",
        database_path: str | Path | None = None,
    ) -> dict[str, Any]:
        context.require_valid()
        AuthorizationService.require_permission(context, PERMISSION_USER_MANAGE_TENANT)

        role_code = AuthorizationService.normalize_role(role)
        location = str(country_location or "").strip()
        decision = AuthorizationService.can_manage_target(
            context,
            target_role=role_code,
            target_country_location=location,
        )
        if not decision.allowed:
            raise PermissionError(decision.reason)

        login_id = AuthService.normalize_user_id(employee_user_id)
        display_name = str(full_name or "").strip()
        email_value = AuthService.normalize_email(email)
        AuthService._validate_identity(login_id, display_name, email_value, location)
        cls._validate_date_range(valid_from, valid_until)
        initial_status = str(account_status or "RESET_REQUIRED").strip().upper()
        if initial_status not in {"RESET_REQUIRED", "DISABLED"}:
            raise ValueError(
                "New accounts must start as RESET_REQUIRED or DISABLED."
            )

        password = (
            PasswordService.generate_temporary_password()
            if generate_password or not str(temporary_password or "")
            else str(temporary_password)
        )
        password_hash, salt, iterations = PasswordService.hash_password(
            password,
            temporary=True,
        )
        expiry = cls._temporary_expiry()

        repository = UserRepository(database_path)
        try:
            created = repository.create_provisioned_user(
                employee_user_id=login_id,
                display_name=display_name,
                email=email_value,
                country_location=location,
                role_code=role_code,
                password_hash=password_hash,
                password_salt=salt,
                password_iterations=iterations,
                account_status=initial_status,
                must_change_password=True,
                temporary_password_expires_at=expiry,
                time_zone=time_zone,
                department=department,
                business_unit=business_unit,
                manager_user_id=manager_user_id,
                valid_from=valid_from,
                valid_until=valid_until,
                created_by_user_id=context.user_id,
            )
            repository.audit_event(
                tenant_id=context.tenant_id,
                actor_user_id=context.user_id,
                action="USER_CREATED",
                target_type="user",
                target_id=str(created["user_id"]),
                details={
                    "employee_user_id": login_id,
                    "role": role_code,
                    "country_location": location,
                    "account_status": initial_status,
                },
            )
            return {
                **created,
                "temporary_password": password,
                "temporary_password_expires_at": expiry,
            }
        finally:
            repository.close()

    @classmethod
    def list_users(
        cls,
        context: SecurityContext,
        database_path: str | Path | None = None,
    ) -> list[dict[str, Any]]:
        AuthorizationService.require_permission(context, PERMISSION_USER_MANAGE_TENANT)
        repository = UserRepository(database_path)
        try:
            country_scope = (
                context.country_location
                if context.role == TENANT_ADMIN
                else None
            )
            return repository.list_users(country_location=country_scope)
        finally:
            repository.close()

    @classmethod
    def reset_temporary_password(
        cls,
        context: SecurityContext,
        *,
        target_database_user_id: int,
        temporary_password: str = "",
        generate_password: bool = False,
        database_path: str | Path | None = None,
    ) -> dict[str, Any]:
        AuthorizationService.require_permission(context, PERMISSION_USER_MANAGE_TENANT)
        repository = UserRepository(database_path)
        try:
            target = repository.get_user_by_id(int(target_database_user_id))
            if not target:
                raise LookupError("The selected user was not found.")
            decision = AuthorizationService.can_manage_target(
                context,
                target_role=str(target.get("role_code") or "USER"),
                target_country_location=str(target.get("country_location") or ""),
            )
            if not decision.allowed:
                raise PermissionError(decision.reason)

            password = (
                PasswordService.generate_temporary_password()
                if generate_password or not str(temporary_password or "")
                else str(temporary_password)
            )
            password_hash, salt, iterations = PasswordService.hash_password(
                password,
                temporary=True,
            )
            expiry = cls._temporary_expiry()
            repository.update_password(
                int(target["id"]),
                password_hash=password_hash,
                password_salt=salt,
                password_iterations=iterations,
                account_status="RESET_REQUIRED",
                must_change_password=True,
                temporary_password_expires_at=expiry,
                updated_by_user_id=context.user_id,
            )
            repository.revoke_all_sessions_for_user(int(target["id"]))
            repository.resolve_password_reset_requests(int(target["id"]), context.user_id)
            repository.audit_event(
                tenant_id=context.tenant_id,
                actor_user_id=context.user_id,
                action="TEMPORARY_CREDENTIAL_RESET",
                target_type="user",
                target_id=str(target["id"]),
                details={"employee_user_id": str(target.get("employee_user_id") or "")},
            )
            return {
                "user_id": str(target.get("employee_user_id") or ""),
                "full_name": str(target.get("display_name") or ""),
                "email": str(target.get("email") or ""),
                "role": str(target.get("role_code") or ""),
                "country_location": str(target.get("country_location") or ""),
                "temporary_password": password,
                "temporary_password_expires_at": expiry,
            }
        finally:
            repository.close()

    @classmethod
    def change_role(
        cls,
        context: SecurityContext,
        *,
        target_database_user_id: int,
        new_role: str,
        database_path: str | Path | None = None,
    ) -> None:
        repository = UserRepository(database_path)
        try:
            target = repository.get_user_by_id(target_database_user_id)
            if not target:
                raise LookupError("The selected user was not found.")
            decision = AuthorizationService.can_manage_target(
                context,
                target_role=new_role,
                target_country_location=str(target.get("country_location") or ""),
            )
            if not decision.allowed:
                raise PermissionError(decision.reason)
            old_role = str(target.get("role_code") or "")
            repository.update_user_role(
                target_database_user_id,
                role_code=AuthorizationService.normalize_role(new_role),
                assigned_by_user_id=context.user_id,
            )
            repository.revoke_all_sessions_for_user(target_database_user_id)
            repository.audit_event(
                tenant_id=context.tenant_id,
                actor_user_id=context.user_id,
                action="USER_ROLE_CHANGED",
                target_type="user",
                target_id=str(target_database_user_id),
                details={"old_role": old_role, "new_role": new_role},
            )
        finally:
            repository.close()

    @classmethod
    def update_account_status(
        cls,
        context: SecurityContext,
        *,
        target_database_user_id: int,
        account_status: str,
        database_path: str | Path | None = None,
    ) -> None:
        repository = UserRepository(database_path)
        try:
            target = repository.get_user_by_id(target_database_user_id)
            if not target:
                raise LookupError("The selected user was not found.")
            decision = AuthorizationService.can_manage_target(
                context,
                target_role=str(target.get("role_code") or "USER"),
                target_country_location=str(target.get("country_location") or ""),
            )
            if not decision.allowed:
                raise PermissionError(decision.reason)
            repository.update_account_status(
                target_database_user_id,
                account_status=account_status,
                updated_by_user_id=context.user_id,
            )
            repository.audit_event(
                tenant_id=context.tenant_id,
                actor_user_id=context.user_id,
                action="USER_STATUS_CHANGED",
                target_type="user",
                target_id=str(target_database_user_id),
                details={"account_status": str(account_status).upper()},
            )
        finally:
            repository.close()

    # ------------------------------------------------------------------
    # Excel import and exports
    # ------------------------------------------------------------------

    @classmethod
    def build_import_template(cls) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Users"
        for column, header in enumerate(cls.IMPORT_COLUMNS, start=1):
            cell = sheet.cell(1, column, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=ALTEN_NAVY.replace("#", ""))
            cell.alignment = Alignment(horizontal="center")
            sheet.column_dimensions[get_column_letter(column)].width = max(18, len(header) + 3)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(cls.IMPORT_COLUMNS))}1"

        instructions = workbook.create_sheet("Instructions")
        instructions.append(["Field", "Requirement"])
        instructions.append(["User ID", "Required; globally unique; stored as text"])
        instructions.append(["Full Name", "Required"])
        instructions.append(["Email", "Required; globally unique"])
        instructions.append(["Role", "SYSTEM_OWNER is not importable; use GLOBAL_ADMIN, TENANT_ADMIN, USER or READER as permitted"])
        instructions.append(["Country or Location", "Required profile and Tenant Admin scope"])
        instructions.append(["Temporary Password", "Optional; blank generates a unique value; plaintext is not stored"])
        instructions.append(["Time Zone", "Optional"])
        for cell in instructions[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=ALTEN_BLUE.replace("#", ""))
        instructions.column_dimensions["A"].width = 26
        instructions.column_dimensions["B"].width = 100

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @classmethod
    def preview_import(
        cls,
        context: SecurityContext,
        workbook_bytes: bytes,
        *,
        database_path: str | Path | None = None,
    ) -> dict[str, Any]:
        AuthorizationService.require_permission(context, PERMISSION_USER_MANAGE_TENANT)
        if not workbook_bytes:
            raise ValueError("Upload a user-import Excel workbook.")
        try:
            dataframe = pd.read_excel(BytesIO(workbook_bytes), sheet_name="Users", dtype=str)
        except ValueError:
            dataframe = pd.read_excel(BytesIO(workbook_bytes), dtype=str)
        dataframe = dataframe.fillna("")
        missing = sorted(cls.REQUIRED_COLUMNS - set(dataframe.columns))
        if missing:
            raise ValueError("Missing required columns: " + ", ".join(missing))

        repository = UserRepository(database_path)
        try:
            seen: set[str] = set()
            rows: list[dict[str, Any]] = []
            for excel_index, source in dataframe.iterrows():
                row_number = int(excel_index) + 2
                item = {column: str(source.get(column, "") or "").strip() for column in cls.IMPORT_COLUMNS}
                errors: list[str] = []
                login_id = item["User ID"]
                role = str(item["Role"] or "USER").upper().replace(" ", "_")
                location = item["Country or Location"]
                try:
                    AuthService._validate_identity(
                        login_id,
                        item["Full Name"],
                        AuthService.normalize_email(item["Email"]),
                        location,
                    )
                except ValueError as exc:
                    errors.append(str(exc))
                normalized = login_id.casefold()
                if normalized in seen:
                    errors.append("Duplicate User ID inside the workbook.")
                seen.add(normalized)
                if repository.get_user_by_login_id(login_id):
                    errors.append("User ID already exists in RecruitOS.")
                try:
                    role = AuthorizationService.normalize_role(role)
                    decision = AuthorizationService.can_manage_target(
                        context,
                        target_role=role,
                        target_country_location=location,
                    )
                    if not decision.allowed:
                        errors.append(decision.reason)
                except ValueError as exc:
                    errors.append(str(exc))
                account_status = str(item["Account Status"] or "").strip().upper()
                if account_status not in {"RESET_REQUIRED", "DISABLED"}:
                    errors.append(
                        "Account Status must be RESET_REQUIRED or DISABLED for a new user."
                    )
                if item["Temporary Password"]:
                    try:
                        PasswordService.validate_temporary_password(item["Temporary Password"])
                    except ValueError as exc:
                        errors.append(str(exc))
                try:
                    cls._validate_date_range(item["Valid From"], item["Valid Until"])
                except ValueError as exc:
                    errors.append(str(exc))

                rows.append(
                    {
                        "row_number": row_number,
                        "user_id": login_id,
                        "full_name": item["Full Name"],
                        "email": AuthService.normalize_email(item["Email"]),
                        "role": role,
                        "country_location": location,
                        "account_status": account_status,
                        "time_zone": item["Time Zone"],
                        "department": item["Department"],
                        "business_unit": item["Business Unit"],
                        "manager_user_id": item["Manager User ID"],
                        "valid_from": item["Valid From"],
                        "valid_until": item["Valid Until"],
                        "temporary_password": item["Temporary Password"],
                        "remarks": item["Remarks"],
                        "valid": not errors,
                        "errors": errors,
                    }
                )
        finally:
            repository.close()

        valid_count = sum(1 for row in rows if row["valid"])
        return {
            "rows": rows,
            "summary": {
                "total": len(rows),
                "valid": valid_count,
                "invalid": len(rows) - valid_count,
            },
        }

    @classmethod
    def commit_import(
        cls,
        context: SecurityContext,
        preview: dict[str, Any],
        *,
        filename: str,
        database_path: str | Path | None = None,
    ) -> dict[str, Any]:
        credentials: list[dict[str, Any]] = []
        results: list[dict[str, Any]] = []
        for row in list(preview.get("rows") or []):
            if not row.get("valid"):
                results.append(cls._safe_import_result(row, outcome="FAILED"))
                continue
            try:
                created = cls.create_user(
                    context,
                    employee_user_id=row["user_id"],
                    full_name=row["full_name"],
                    email=row["email"],
                    role=row["role"],
                    country_location=row["country_location"],
                    temporary_password=row["temporary_password"],
                    generate_password=not bool(row["temporary_password"]),
                    account_status=row["account_status"],
                    time_zone=row["time_zone"],
                    department=row["department"],
                    business_unit=row["business_unit"],
                    manager_user_id=row["manager_user_id"],
                    valid_from=row["valid_from"],
                    valid_until=row["valid_until"],
                    database_path=database_path,
                )
                credential = cls._credential_record(created)
                credentials.append(credential)
                results.append(cls._safe_import_result(row, outcome="CREATED", errors=[]))
            except Exception as exc:  # row-level isolation is deliberate
                results.append(
                    cls._safe_import_result(row, outcome="FAILED", errors=[str(exc)])
                )

        created_count = sum(1 for item in results if item["outcome"] == "CREATED")
        failed_count = len(results) - created_count
        repository = UserRepository(database_path)
        try:
            repository.record_import_job(
                uploaded_by_user_id=context.user_id,
                filename=filename,
                total_rows=len(results),
                created_rows=created_count,
                skipped_rows=0,
                failed_rows=failed_count,
                summary={"created": created_count, "failed": failed_count},
            )
            repository.audit_event(
                tenant_id=context.tenant_id,
                actor_user_id=context.user_id,
                action="USERS_IMPORTED",
                target_type="user_import",
                target_id=filename,
                details={"created": created_count, "failed": failed_count},
            )
        finally:
            repository.close()
        return {
            "results": results,
            "credentials": credentials,
            "created": created_count,
            "failed": failed_count,
        }

    @classmethod
    def export_user_access_master(
        cls,
        context: SecurityContext,
        database_path: str | Path | None = None,
    ) -> bytes:
        AuthorizationService.require_permission(context, PERMISSION_ACCESS_MASTER)
        rows = cls.list_users(context, database_path)
        headers = [
            "User ID",
            "Full Name",
            "Email",
            "Role",
            "Country or Location",
            "Account Status",
            "Time Zone",
            "Department",
            "Business Unit",
            "Manager User ID",
            "Must Change Password",
            "Temporary Password Expiry",
            "Valid From",
            "Valid Until",
            "Last Login",
            "Failed Login Attempts",
            "Locked Until",
            "Created Date",
            "Updated Date",
        ]
        data = [
            [
                row.get("employee_user_id", ""),
                row.get("display_name", ""),
                row.get("email", ""),
                ROLE_LABELS.get(str(row.get("role_code") or ""), row.get("role_code", "")),
                row.get("country_location", ""),
                row.get("account_status", ""),
                row.get("time_zone", ""),
                row.get("department", ""),
                row.get("business_unit", ""),
                row.get("manager_user_id", ""),
                "Yes" if row.get("must_change_password") else "No",
                row.get("temporary_password_expires_at", ""),
                row.get("valid_from", ""),
                row.get("valid_until", ""),
                row.get("last_login_at", ""),
                row.get("failed_login_count", 0),
                row.get("locked_until", ""),
                row.get("created_at", ""),
                row.get("updated_at", ""),
            ]
            for row in rows
        ]
        return cls._build_workbook("User Access Master", headers, data)

    @classmethod
    def export_temporary_credentials(cls, credentials: list[dict[str, Any]]) -> bytes:
        headers = [
            "User ID",
            "Full Name",
            "Email",
            "Role",
            "Country or Location",
            "Temporary Password",
            "Expiry",
            "First Login Reset Required",
        ]
        data = [
            [
                item.get("user_id", ""),
                item.get("full_name", ""),
                item.get("email", ""),
                item.get("role", ""),
                item.get("country_location", ""),
                item.get("temporary_password", ""),
                item.get("temporary_password_expires_at", ""),
                "Yes",
            ]
            for item in credentials
        ]
        return cls._build_workbook("Temporary Credentials", headers, data)

    @classmethod
    def pending_reset_requests(
        cls,
        context: SecurityContext,
        database_path: str | Path | None = None,
    ) -> list[dict[str, Any]]:
        AuthorizationService.require_permission(context, PERMISSION_USER_MANAGE_TENANT)
        repository = UserRepository(database_path)
        try:
            scope = context.country_location if context.role == TENANT_ADMIN else None
            return repository.list_pending_password_reset_requests(country_location=scope)
        finally:
            repository.close()

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _temporary_expiry() -> str:
        if AUTH_TEMP_PASSWORD_EXPIRY_DAYS <= 0:
            return ""
        return (
            datetime.now(timezone.utc)
            + timedelta(days=AUTH_TEMP_PASSWORD_EXPIRY_DAYS)
        ).isoformat(timespec="seconds")

    @staticmethod
    def _validate_date_range(valid_from: str, valid_until: str) -> None:
        start_text = str(valid_from or "").strip()
        end_text = str(valid_until or "").strip()
        if not start_text or not end_text:
            return
        try:
            start = datetime.fromisoformat(start_text)
            end = datetime.fromisoformat(end_text)
        except ValueError as exc:
            raise ValueError("Valid From and Valid Until must be valid dates.") from exc
        if end < start:
            raise ValueError("Valid Until must be later than Valid From.")

    @staticmethod
    def _credential_record(created: dict[str, Any]) -> dict[str, Any]:
        return {
            "user_id": created.get("employee_user_id", ""),
            "full_name": created.get("display_name", ""),
            "email": created.get("email", ""),
            "role": created.get("role", ""),
            "country_location": created.get("country_location", ""),
            "temporary_password": created.get("temporary_password", ""),
            "temporary_password_expires_at": created.get("temporary_password_expires_at", ""),
        }

    @staticmethod
    def _safe_import_result(
        row: dict[str, Any],
        *,
        outcome: str,
        errors: list[str] | None = None,
    ) -> dict[str, Any]:
        """Return an import result with plaintext credentials removed."""
        safe = {
            key: value
            for key, value in row.items()
            if key != "temporary_password"
        }
        safe["outcome"] = str(outcome)
        if errors is not None:
            safe["errors"] = list(errors)
        return safe

    @staticmethod
    def _build_workbook(title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = title[:31]
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(1, column, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=ALTEN_NAVY.replace("#", ""))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.column_dimensions[get_column_letter(column)].width = max(16, min(34, len(header) + 4))
        for row_index, values in enumerate(rows, start=2):
            for column, value in enumerate(values, start=1):
                sheet.cell(row_index, column, value)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()
